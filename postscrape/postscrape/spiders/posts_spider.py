import scrapy
import requests
from bs4 import BeautifulSoup
from datetime import datetime, date, timedelta
import xlrd
from xlwt import Workbook 
import plotly.graph_objects as go
import pandas as pd
from openpyxl import load_workbook
from time import sleep
import os
import numpy as np
from statistics import mean




class PostsSpider(scrapy.Spider):
    name = 'posts'
    #######begin bagian yang diubah ubah
    partisi = 2
    #######end bagian yang diubah ubah
    start_urls = []
    list_emiten_used = []
    itung = 0    
    baris_analyst = 1
    baris = 1
    sebelum = datetime.now()
    sesudah = datetime.now()
    wb_analyst = Workbook()
    sheet_analyst = ''
    pd_data = []
    list_jii70 = []

    def __init__(self):
        print('init dimulai')
        self.sebelum = datetime.now()
        self.isi_urls()
        self.delete_files()
        self.excel_initialization()

    def parse(self, response):
        self.pd_data = []
        self.itung = self.itung + 1
        print('ini itungan ke = '+str(self.itung))
        print(response.url)
        emiten = response.url.split('/')[-2].split('.')[0]
        src = response.text
        rows = self.get_table_row(src)
        if len(rows) < 90:
            src = requests.get(response.url).text
            rows = self.get_table_row(src)

        if len(rows) >= 90:
            self.rows_process(rows, emiten)


        if self.itung == len(self.start_urls):
            self.wb_analyst.save('D:\BelajarData\dataset\saham\JII_analysis.xlsx')
            from matplotlib.pyplot import savefig
            import mplfinance as mpf
            directory = 'D:\BelajarData\dataset\saham\emiten'
            pic_directory = 'D:\BelajarData\dataset\saham\emiten_pic'
            pic_directory_up = 'D:\BelajarData\dataset\saham\emiten_pic_up'
            count_pic = 1
            stats_data = []
            for filename in os.listdir(directory):
                file_name = f'{directory}\{filename}'
                df_saham = pd.read_excel(file_name, engine='openpyxl')
                stat_data = self.stat_process(df_saham)
                stats_data.append(stat_data)
                if df_saham['JII'][0] in self.list_emiten_used:
                    if stat_data[2] == 'up_near' or stat_data[1] == 'up' or stat_data[3] == 'low':
                        df_saham = df_saham.reindex(index=df_saham.index[::-1])
                        df_saham = df_saham.set_index(pd.DatetimeIndex(df_saham['date'].values))
                        emiten = filename.split('.')[0]
                        print(emiten, count_pic)
                        pic_file_name = f'{pic_directory_up}\{emiten}.jpg'
                        save = dict(fname=pic_file_name,dpi=100)
                        mpf.plot(df_saham,type='candle',mav=(3,7),volume=True,figsize=(25,15),savefig=save,title=emiten)
                        count_pic += 1                    
                    elif df_saham['JII'][0] in self.list_jii70:
                        df_saham = df_saham.reindex(index=df_saham.index[::-1])
                        df_saham = df_saham.set_index(pd.DatetimeIndex(df_saham['date'].values))
                        emiten = filename.split('.')[0]
                        print(emiten, count_pic)
                        pic_file_name = f'{pic_directory}\{emiten}.jpg'
                        save = dict(fname=pic_file_name,dpi=100)
                        mpf.plot(df_saham,type='candle',mav=(3,7),volume=True,figsize=(25,15),savefig=save,title=emiten)
                        count_pic += 1
            df_stat = pd.DataFrame(stats_data,
                                    columns=['emiten', 'macd', 'up_near_macd','get_low', 'rsi','percent_volum',
                                            'Open','High','Low','Close','Volume'])
            df_stat.to_excel('D:\BelajarData\dataset\saham\daily_stock_analysis_v3.xlsx')
            self.sesudah = datetime.now()
            selisih = self.sesudah - self.sebelum
            print(selisih)
            print(f'selesai pakai partisi {self.partisi}')


    def stat_process(self, df):
        emiten = df['JII'][0]
        last_open = df['Open'][0]
        last_high = df['High'][0]
        last_low = df['Low'][0]
        last_close = df['Close'][0]
        last_volume = df['Volume'][0]
        macd = self.calculate_macd(df['Close'])
        up_near_macd = self.calculate_up_near_macd(df['Close'])
        rsi = self.calculate_rsi(df['Close'])
        percent_volum = self.calculate_percent_volum(df['Volume'])
        get_low = self.get_low(df)
        print(df['date'][0])
        return [emiten, macd, up_near_macd, get_low, rsi, percent_volum, last_open, last_high,
        last_low,last_close,last_volume]

    def get_low(self, df):
        current_close = df['Close'][0]
        previous_close = df['Close'][1]
        current_open = df['Open'][0]
        # current_high = df['High'][0]
        rsi5 = self.calculate_rsi(df['Close'], rsi_coverage=5)
        con_a = current_close >= previous_close
        con_b = current_close < previous_close
        con_c = current_close > current_open
        con_d = con_b and con_c
        # con_e = con_b or (con_c and con_d)
        con_f = rsi5 <= 30
        if (con_a or con_d) and con_f:
            return 'low'
        return '---'

    def calculate_macd(self, df_close, ma_small=3, ma_big=7):
        current_ma_small = mean(df_close[:ma_small])
        previous_ma_small = mean(df_close[1:ma_small+1])
        current_ma_big = mean(df_close[:ma_big])
        previous_ma_big = mean(df_close[1:ma_big+1])

        con_a = current_ma_small > previous_ma_small
        con_b = current_ma_small >= current_ma_big
        con_c = previous_ma_small < previous_ma_big

        con_d = current_ma_small < previous_ma_small
        con_e = current_ma_small <= current_ma_big
        con_f = previous_ma_small > previous_ma_big

        if con_a and con_b and con_c:
            return 'up'
        elif con_d and con_e and con_f:
            return 'down'
        else:
            return '---'

    def calculate_up_near_macd(self, df_close, near_degree=0.5, ma_small=3, ma_big=7):
        current_ma_small = mean(df_close[:ma_small])
        previous_ma_small = mean(df_close[1:ma_small+1])
        current_ma_big = mean(df_close[:ma_big])
        previous_ma_big = mean(df_close[1:ma_big+1])

        con_a = current_ma_big >= current_ma_small
        con_b = current_ma_small > previous_ma_small
        con_c = (current_ma_big-current_ma_small) < (previous_ma_big-previous_ma_small)
        con_d = previous_ma_big > previous_ma_small
        con_e = current_ma_big <= previous_ma_big

        if con_a and con_b and con_c and con_d and con_e:
            x = previous_ma_big - previous_ma_small - current_ma_big + current_ma_small
            if x != 0:
                distance = (current_ma_big-current_ma_small)/x
                if distance >=0 and distance <= near_degree:
                    return 'up_near'
        return '---'
        

    def calculate_percent_volum(self, df_volume, coverage=7):
        current_volume = df_volume[0]
        mean_volume = mean(df_volume[:7])
        percent_volume = (current_volume-mean_volume)*100/mean_volume
        return percent_volume

    def calculate_rsi(self, df_close, rsi_coverage=14):
        cssdh = np.asarray(df_close[:-1])
        csblm = np.asarray(df_close[1:])
        cselisih = cssdh - csblm
        cup = np.copy(cselisih)
        cdown = np.copy(cselisih)
        cup[cup<0]=0
        cdown[cdown>0]=0
        cdown = cdown * -1
        if max(cup[:rsi_coverage]) == 0:
            up_value = 0.001
        else:
            up_value = np.mean(cup[:rsi_coverage])

        if max(cdown[:rsi_coverage]) == 0:
            down_value = 0.001
        else:
            down_value = np.mean(cdown[:rsi_coverage])

        banding = up_value/down_value
        rsi = 100 - (100/(banding+1))
        return rsi

    def get_table_row(self, src):
        soup = BeautifulSoup(src,'lxml')
        table = soup.findAll('table')[0]
        body = table.findAll('tbody')[0]
        rows = body.findAll('tr')
        return rows
    
    def rows_process(self, rows, emiten, baris_emiten = 1):
        for row in rows:
            pd_data_kolom = []
            kolom = row.findAll('td')
            count = 0
            
            for data in kolom:
                if len(kolom) < 6 :                    
                    break

                volume = kolom[6].text
                if volume == '-':
                    break

                volume = kolom[6].span.text
                if len(str(volume)) <= 2:
                    break

                if count > 0 :
                    if count != 5 :
                        value = data.span.text
                        for character in [',','.00']:
                            value = value.replace(character, '')

                        if count > 5:
                            if baris_emiten <= 30:
                                self.sheet_analyst.write(self.baris_analyst, count, int(value)) 
                                self.baris_analyst = self.baris_analyst +1
                            
                            pd_data_kolom.append(value)
                            self.baris = self.baris + 1
                            baris_emiten = baris_emiten + 1
                        else:
                            pd_data_kolom.append(value)
                            if baris_emiten <= 30:
                                self.sheet_analyst.write(self.baris_analyst, count+1, int(value))
                else:
                    value = data.span.text
                    value = datetime.strptime(value, '%b %d, %Y') 
                    value = datetime.strftime(value, '%Y-%m-%d') 
                    pd_data_kolom.append(value)
                    pd_data_kolom.append(emiten)
                    if row == rows[0]:
                        print(value, emiten)

                    if baris_emiten <= 30:
                        self.sheet_analyst.write(self.baris_analyst, 0, value) 
                        self.sheet_analyst.write(self.baris_analyst, 1, emiten) 

                count = count + 1

            if count > 0:
                self.pd_data.append(pd_data_kolom)

            if baris_emiten > 70:
                df_saham = pd.DataFrame(self.pd_data,
                    columns=['date', 'JII', 'Open','High','Low','Close','Volume'])

                self.create_stock_file(df_saham, emiten)
                break

    def create_stock_file(self, df_saham, emiten):
        df_saham.to_excel(f'D:\BelajarData\dataset\saham\emiten\{emiten}.xlsx')
        # df_saham = df_saham.reindex(index=df_saham.index[::-1])
        # df_saham = df_saham.set_index(pd.DatetimeIndex(df_saham['date'].values))
        # figure = go.Figure(
        #         data =[
        #             go.Candlestick(
        #             x = df_saham.index,
        #             low = df_saham['Low'],
        #             high = df_saham['High'],
        #             close = df_saham['Close'],
        #             open = df_saham['Open'],
        #             increasing_line_color = 'green',
        #             decreasing_line_color = 'red',
        #             )
        #         ],
        #         layout=go.Layout(
        #             title=go.layout.Title(text=emiten)
        #         )
        #     )
        # figure.write_image(file = 'D:\BelajarData\dataset\saham\'+emiten+'.png',
        #           width = 1600,
        #           height = 1000)

    def isi_urls(self):
        loc = "D:\BelajarData\dataset\saham\JII.xlsx"
        wb_jii_list = load_workbook(loc)

        sheet_jii_70 = wb_jii_list[wb_jii_list.sheetnames[0]]
        for x in range(sheet_jii_70.max_row):
            emiten = sheet_jii_70[f'A{str(x+1)}'].value
            self.list_jii70.append(emiten)

        sheet_jii_large = wb_jii_list[wb_jii_list.sheetnames[1]]
        # self.start_urls.append('https://finance.yahoo.com/quote/HRUM.JK/history?p=HRUM.JK')
        # self.list_emiten_used.append('HRUM')
        hitung = 0
        for i in range(sheet_jii_large.max_row):
            emiten = sheet_jii_large[f'A{str(i+1)}'].value
            if (hitung%3) == self.partisi:
                self.list_emiten_used.append(emiten)
                self.start_urls.append('https://finance.yahoo.com/quote/'+emiten+'.JK/history?p='+emiten+'.JK')
            hitung = hitung + 1

    def excel_initialization(self):
        self.sheet_analyst = self.wb_analyst.add_sheet('Sheet 1')
        self.sheet_analyst.write(0, 0, 'date') 
        self.sheet_analyst.write(0, 1, 'JII') 
        self.sheet_analyst.write(0, 2, 'Open') 
        self.sheet_analyst.write(0, 3, 'High') 
        self.sheet_analyst.write(0, 4, 'Low') 
        self.sheet_analyst.write(0, 5, 'Close') 
        self.sheet_analyst.write(0, 6, 'Volume')

    def delete_files(self):
        list_dir=[
            'D:\BelajarData\dataset\saham\emiten',
            'D:\BelajarData\dataset\saham\emiten_pic',
            'D:\BelajarData\dataset\saham\emiten_pic_up',
        ]
        for dir in list_dir:
            if len(os.listdir(dir)) > 0:
                for filename in os.listdir(dir):
                    emiten = filename.split('.')[0]
                    file_name = f'{dir}\{filename}'
                    if os.path.isfile(file_name):
                        if emiten in self.list_emiten_used:
                            os.remove(file_name)
